from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import json


class ReportService:
    """
    Service dùng để sinh báo cáo Evaluation.

    Hỗ trợ:
    - Excel (.xlsx)
    - JSON (.json)
    """

    REPORT_FOLDER = (
        Path(__file__).resolve().parents[2]
        / "reports"
    )

    def __init__(self):
        self.REPORT_FOLDER.mkdir(exist_ok=True)

    # ==========================================
    # Excel Report
    # ==========================================

    def generate_excel_report(
        self,
        results,
        model,
        dataset,
        average_score,
        pass_rate
    ):

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        bold = Font(bold=True)

        green_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
        )

        red_fill = PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid"
        )

        # ==========================
        # Summary Sheet
        # ==========================

        summary_data = [
            ("Model", model),
            ("Dataset", dataset),
            ("Average Score", average_score),
            ("Pass Rate (%)", pass_rate),
            ("Generated Time",
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]

        for row in summary_data:
            summary_sheet.append(row)

        for cell in summary_sheet["A"]:
            cell.font = bold

        # ==========================
        # Results Sheet
        # ==========================

        result_sheet = workbook.create_sheet("Results")

        headers = [
            "Question",
            "Expected",
            "Answer",
            "Score",
            "Status",
            "Reason"
        ]

        result_sheet.append(headers)

        for cell in result_sheet[1]:
            cell.font = bold

        for item in results:

            reason = ""

            if "metrics" in item:

                metric_name = next(
                    iter(item["metrics"])
                )

                reason = item["metrics"][metric_name].get(
                    "reason",
                    ""
                )

            result_sheet.append([
                item["question"],
                item["expected"],
                item["answer"],
                item["score"],
                item["status"],
                reason
            ])

            current_row = result_sheet.max_row

            status_cell = result_sheet.cell(
                row=current_row,
                column=5
            )

            if item["status"] == "PASS":
                status_cell.fill = green_fill
            else:
                status_cell.fill = red_fill

        # ==========================
        # Auto width
        # ==========================

        for sheet in workbook.worksheets:

            for column in sheet.columns:

                max_length = 0

                column_letter = get_column_letter(
                    column[0].column
                )

                for cell in column:

                    try:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                    except:
                        pass

                sheet.column_dimensions[
                    column_letter
                ].width = min(max_length + 3, 60)

        filename = (
            "evaluation_report_"
            + datetime.now().strftime("%Y%m%d_%H%M%S")
            + ".xlsx"
        )

        file_path = self.REPORT_FOLDER / filename

        workbook.save(file_path)

        return filename

    # ==========================================
    # JSON Report
    # ==========================================

    def generate_json_report(
        self,
        results,
        model,
        dataset,
        average_score,
        pass_rate
    ):

        filename = (
            "evaluation_report_"
            + datetime.now().strftime("%Y%m%d_%H%M%S")
            + ".json"
        )

        file_path = self.REPORT_FOLDER / filename

        report = {
            "summary": {
                "model": model,
                "dataset": dataset,
                "average_score": average_score,
                "pass_rate": pass_rate,
                "generated_time": datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
            },
            "results": results
        }

        with open(
            file_path,
            "w",
            encoding="utf-8"
        ) as f:

            json.dump(
                report,
                f,
                indent=4,
                ensure_ascii=False
            )

        return filename

    # ==========================================
    # Utils
    # ==========================================

    def get_report_path(self, filename):

        return self.REPORT_FOLDER / filename

    def report_exists(self, filename):

        return self.get_report_path(
            filename
        ).exists()